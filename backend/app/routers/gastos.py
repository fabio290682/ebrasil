import csv
import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..config import DATABASE_BACKEND
from ..database import get_db
from ..schemas import GastoListResponse, PageMeta, ResumoGastosResponse, TopFornecedorItem
from ..services import mongo_query
from ..services.query import list_gastos, resumo_gastos, top_fornecedores

router = APIRouter(prefix="/gastos", tags=["gastos"])

_EXPORT_COLS = [
    ("id", "ID"),
    ("data_empenho", "Data"),
    ("favorecido_nome", "Favorecido"),
    ("favorecido_cnpj_cpf", "CNPJ/CPF"),
    ("valor_empenhado", "Valor (R$)"),
    ("funcao_governo", "Função"),
    ("elemento_despesa", "Elemento"),
    ("fonte_recurso", "Fonte"),
    ("uf", "UF"),
    ("municipio_ibge", "IBGE"),
    ("categoria_origem", "Categoria"),
    ("agente_publico", "Agente"),
    ("partido", "Partido"),
    ("numero_empenho", "Nº Empenho"),
    ("fornecedor_sistema", "Sistema"),
    ("url_origem", "URL"),
]


def _export_params(
    data_inicio: date | None,
    data_fim: date | None,
    uf: str | None,
    municipio_ibge: str | None,
    elemento_despesa: str | None,
    fornecedor: str | None,
    categoria_origem: str | None,
    agente_publico: str | None,
    partido: str | None,
    db: Session,
):
    if DATABASE_BACKEND == "mongo":
        items, _, _ = mongo_query.list_gastos(
            page=1, page_size=10000,
            data_inicio=data_inicio, data_fim=data_fim,
            uf=uf, municipio_ibge=municipio_ibge,
            elemento_despesa=elemento_despesa, fornecedor=fornecedor,
            categoria_origem=categoria_origem, agente_publico=agente_publico,
            partido=partido,
        )
    else:
        items, _, _ = list_gastos(
            db, page=1, page_size=10000,
            data_inicio=data_inicio, data_fim=data_fim,
            uf=uf, municipio_ibge=municipio_ibge,
            elemento_despesa=elemento_despesa, fornecedor=fornecedor,
            categoria_origem=categoria_origem, agente_publico=agente_publico,
            partido=partido,
        )
    return items


def _getv(obj, key):
    if isinstance(obj, dict):
        return obj.get(key)
    return getattr(obj, key, None)


@router.get("", response_model=GastoListResponse)
def get_gastos(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    data_inicio: date | None = Query(default=None),
    data_fim: date | None = Query(default=None),
    uf: str | None = Query(default=None, min_length=2, max_length=2),
    municipio_ibge: str | None = Query(default=None, min_length=7, max_length=7),
    elemento_despesa: str | None = Query(default=None),
    fornecedor: str | None = Query(default=None),
    categoria_origem: str | None = Query(default=None),
    agente_publico: str | None = Query(default=None),
    partido: str | None = Query(default=None, min_length=2, max_length=20),
    db: Session = Depends(get_db),
):
    if DATABASE_BACKEND == "mongo":
        items, total, total_pages = mongo_query.list_gastos(
            page=page,
            page_size=page_size,
            data_inicio=data_inicio,
            data_fim=data_fim,
            uf=uf,
            municipio_ibge=municipio_ibge,
            elemento_despesa=elemento_despesa,
            fornecedor=fornecedor,
            categoria_origem=categoria_origem,
            agente_publico=agente_publico,
            partido=partido,
        )
    else:
        items, total, total_pages = list_gastos(
            db,
            page=page,
            page_size=page_size,
            data_inicio=data_inicio,
            data_fim=data_fim,
            uf=uf,
            municipio_ibge=municipio_ibge,
            elemento_despesa=elemento_despesa,
            fornecedor=fornecedor,
            categoria_origem=categoria_origem,
            agente_publico=agente_publico,
            partido=partido,
        )

    return GastoListResponse(
        items=items,
        meta=PageMeta(page=page, page_size=page_size, total=total, total_pages=total_pages),
    )


@router.get("/resumo", response_model=ResumoGastosResponse)
def get_resumo(
    data_inicio: date | None = Query(default=None),
    data_fim: date | None = Query(default=None),
    uf: str | None = Query(default=None, min_length=2, max_length=2),
    municipio_ibge: str | None = Query(default=None, min_length=7, max_length=7),
    categoria_origem: str | None = Query(default=None),
    agente_publico: str | None = Query(default=None),
    partido: str | None = Query(default=None, min_length=2, max_length=20),
    db: Session = Depends(get_db),
):
    if DATABASE_BACKEND == "mongo":
        total, quantidade, ticket = mongo_query.resumo_gastos(
            data_inicio=data_inicio,
            data_fim=data_fim,
            uf=uf,
            municipio_ibge=municipio_ibge,
            categoria_origem=categoria_origem,
            agente_publico=agente_publico,
            partido=partido,
        )
    else:
        total, quantidade, ticket = resumo_gastos(
            db,
            data_inicio=data_inicio,
            data_fim=data_fim,
            uf=uf,
            municipio_ibge=municipio_ibge,
            categoria_origem=categoria_origem,
            agente_publico=agente_publico,
            partido=partido,
        )

    return ResumoGastosResponse(total_empenhado=total, quantidade_registros=quantidade, ticket_medio=ticket)


@router.get("/top-fornecedores", response_model=list[TopFornecedorItem])
def get_top_fornecedores(
    limit: int = Query(default=10, ge=1, le=100),
    data_inicio: date | None = Query(default=None),
    data_fim: date | None = Query(default=None),
    uf: str | None = Query(default=None, min_length=2, max_length=2),
    categoria_origem: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    if DATABASE_BACKEND == "mongo":
        rows = mongo_query.top_fornecedores(
            limit=limit,
            data_inicio=data_inicio,
            data_fim=data_fim,
            uf=uf,
            categoria_origem=categoria_origem,
        )
        return [
            TopFornecedorItem(
                favorecido_nome=r.get("_id") or "NAO INFORMADO",
                total_empenhado=float(r.get("total_empenhado") or 0.0),
                quantidade_empenhos=int(r.get("quantidade_empenhos") or 0),
            )
            for r in rows
        ]

    rows = top_fornecedores(
        db,
        limit=limit,
        data_inicio=data_inicio,
        data_fim=data_fim,
        uf=uf,
        categoria_origem=categoria_origem,
    )
    return [
        TopFornecedorItem(
            favorecido_nome=row[0],
            total_empenhado=float(row[1] or 0.0),
            quantidade_empenhos=int(row[2] or 0),
        )
        for row in rows
    ]


@router.get("/export/csv")
def export_csv(
    data_inicio: date | None = Query(default=None),
    data_fim: date | None = Query(default=None),
    uf: str | None = Query(default=None, min_length=2, max_length=2),
    municipio_ibge: str | None = Query(default=None, min_length=7, max_length=7),
    elemento_despesa: str | None = Query(default=None),
    fornecedor: str | None = Query(default=None),
    categoria_origem: str | None = Query(default=None),
    agente_publico: str | None = Query(default=None),
    partido: str | None = Query(default=None, min_length=2, max_length=20),
    db: Session = Depends(get_db),
):
    if DATABASE_BACKEND == "mongo":
        items, _, _ = mongo_query.list_gastos(
            page=1,
            page_size=10000,
            data_inicio=data_inicio,
            data_fim=data_fim,
            uf=uf,
            municipio_ibge=municipio_ibge,
            elemento_despesa=elemento_despesa,
            fornecedor=fornecedor,
            categoria_origem=categoria_origem,
            agente_publico=agente_publico,
            partido=partido,
        )
    else:
        items, _, _ = list_gastos(
            db,
            page=1,
            page_size=10000,
            data_inicio=data_inicio,
            data_fim=data_fim,
            uf=uf,
            municipio_ibge=municipio_ibge,
            elemento_despesa=elemento_despesa,
            fornecedor=fornecedor,
            categoria_origem=categoria_origem,
            agente_publico=agente_publico,
            partido=partido,
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id",
        "data_empenho",
        "favorecido_nome",
        "favorecido_cnpj_cpf",
        "valor_empenhado",
        "funcao_governo",
        "elemento_despesa",
        "fonte_recurso",
        "uf",
        "municipio_ibge",
        "categoria_origem",
        "agente_publico",
        "partido",
        "numero_empenho",
        "fornecedor_sistema",
        "url_origem",
    ])
    for g in items:
        writer.writerow([
            _getv(g, "id"),
            _getv(g, "data_empenho"),
            _getv(g, "favorecido_nome"),
            _getv(g, "favorecido_cnpj_cpf"),
            _getv(g, "valor_empenhado"),
            _getv(g, "funcao_governo"),
            _getv(g, "elemento_despesa"),
            _getv(g, "fonte_recurso"),
            _getv(g, "uf"),
            _getv(g, "municipio_ibge"),
            _getv(g, "categoria_origem"),
            _getv(g, "agente_publico"),
            _getv(g, "partido"),
            _getv(g, "numero_empenho"),
            _getv(g, "fornecedor_sistema"),
            _getv(g, "url_origem"),
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=gastos_iiibrasil.csv"},
    )


@router.get("/export/xlsx")
def export_xlsx(
    data_inicio: date | None = Query(default=None),
    data_fim: date | None = Query(default=None),
    uf: str | None = Query(default=None, min_length=2, max_length=2),
    municipio_ibge: str | None = Query(default=None, min_length=7, max_length=7),
    elemento_despesa: str | None = Query(default=None),
    fornecedor: str | None = Query(default=None),
    categoria_origem: str | None = Query(default=None),
    agente_publico: str | None = Query(default=None),
    partido: str | None = Query(default=None, min_length=2, max_length=20),
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    items = _export_params(
        data_inicio, data_fim, uf, municipio_ibge,
        elemento_despesa, fornecedor, categoria_origem, agente_publico, partido, db,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos Publicos"

    # Cabeçalho
    header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    headers = [label for _, label in _EXPORT_COLS]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dados
    alt_fill = PatternFill(fill_type="solid", fgColor="F0F4FA")
    money_fmt = '#,##0.00'
    for row_idx, g in enumerate(items, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, (field, _) in enumerate(_EXPORT_COLS, start=1):
            val = _getv(g, field)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            if field == "valor_empenhado" and val is not None:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

    # Larguras aproximadas
    col_widths = [36, 12, 36, 18, 16, 20, 12, 12, 5, 10, 22, 28, 8, 18, 18, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    ts = datetime.now().strftime("%Y%m%d")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gastos_iiibrasil_{ts}.xlsx"},
    )


@router.get("/export/pdf")
def export_pdf(
    data_inicio: date | None = Query(default=None),
    data_fim: date | None = Query(default=None),
    uf: str | None = Query(default=None, min_length=2, max_length=2),
    municipio_ibge: str | None = Query(default=None, min_length=7, max_length=7),
    elemento_despesa: str | None = Query(default=None),
    fornecedor: str | None = Query(default=None),
    categoria_origem: str | None = Query(default=None),
    agente_publico: str | None = Query(default=None),
    partido: str | None = Query(default=None, min_length=2, max_length=20),
    db: Session = Depends(get_db),
):
    from fpdf import FPDF

    items = _export_params(
        data_inicio, data_fim, uf, municipio_ibge,
        elemento_despesa, fornecedor, categoria_origem, agente_publico, partido, db,
    )
    items = items[:2000]  # PDF limitado a 2000 linhas para não travar

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    # Título
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(30, 58, 95)
    pdf.cell(0, 8, "IIIbrasil - Gastos Publicos", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | {len(items)} registros", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Colunas visíveis no PDF (subconjunto legível em A4 landscape)
    pdf_cols = [
        ("data_empenho",     "Data",        22),
        ("favorecido_nome",  "Favorecido",  65),
        ("valor_empenhado",  "Valor (R$)",  30),
        ("funcao_governo",   "Funcao",      30),
        ("uf",               "UF",           9),
        ("categoria_origem", "Categoria",   35),
        ("agente_publico",   "Agente",      40),
        ("partido",          "Partido",     15),
        ("numero_empenho",   "Empenho",     25),
    ]

    # Cabeçalho da tabela
    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 7)
    row_h = 6
    for _, label, w in pdf_cols:
        pdf.cell(w, row_h, label, border=0, fill=True, align="C")
    pdf.ln()

    # Linhas
    pdf.set_font("Helvetica", "", 7)
    for i, g in enumerate(items):
        if pdf.get_y() > 190:
            pdf.add_page()
            pdf.set_fill_color(30, 58, 95)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 7)
            for _, label, w in pdf_cols:
                pdf.cell(w, row_h, label, border=0, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)

        fill = i % 2 == 0
        pdf.set_fill_color(240, 244, 250) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 30, 30)

        for field, _, w in pdf_cols:
            val = _getv(g, field)
            if field == "valor_empenhado" and val is not None:
                text = f"R$ {float(val):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            elif val is None:
                text = ""
            else:
                text = str(val)
            # Truncar texto longo
            pdf.cell(w, row_h, text[:30] if len(text) > 30 else text, border=0, fill=fill)
        pdf.ln()

    ts = datetime.now().strftime("%Y%m%d")
    raw = bytes(pdf.output())
    return StreamingResponse(
        iter([raw]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=gastos_iiibrasil_{ts}.pdf"},
    )
